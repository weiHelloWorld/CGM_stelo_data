
from __future__ import annotations

import argparse
import csv
import math
import zipfile
import xml.etree.ElementTree as ET
from bisect import bisect_left, bisect_right
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager
# _CJK_FONT = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
# font_manager.fontManager.addfont(_CJK_FONT)
# plt.rcParams["font.family"] = font_manager.FontProperties(fname=_CJK_FONT).get_name()
# plt.rcParams["axes.unicode_minus"] = False
try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl。请先运行: pip install openpyxl"
    ) from exc


def col_to_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch.upper()) - ord("A") + 1
    return n - 1


def read_first_sheet_xlsx(path: str) -> list[list[object]]:
    """Read values from the first worksheet using only the Python standard library."""
    ns_main = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(path) as zf:
        shared_strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", ns_main):
                text = "".join(t.text or "" for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
                shared_strings.append(text)

        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        sheet = workbook.find("m:sheets/m:sheet", ns_main)
        if sheet is None:
            raise ValueError("Excel 文件中没有工作表。")
        rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]

        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = None
        for rel in rels:
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib["Target"]
                break
        if target is None:
            raise ValueError("无法解析第一个工作表。")
        sheet_path = target.lstrip("/") if target.startswith("/") else "xl/" + target.lstrip("../")
        root = ET.fromstring(zf.read(sheet_path))

        rows = []
        for row_el in root.findall(".//m:sheetData/m:row", ns_main):
            row_values = []
            for c in row_el.findall("m:c", ns_main):
                idx = col_to_index(c.attrib.get("r", "A1"))
                while len(row_values) <= idx:
                    row_values.append(None)
                cell_type = c.attrib.get("t")
                v = c.find("m:v", ns_main)
                inline = c.find("m:is", ns_main)
                value = None
                if cell_type == "s" and v is not None:
                    value = shared_strings[int(v.text)]
                elif cell_type == "inlineStr" and inline is not None:
                    value = "".join(t.text or "" for t in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
                elif cell_type == "b" and v is not None:
                    value = v.text == "1"
                elif v is not None:
                    raw = v.text or ""
                    try:
                        num = float(raw)
                        value = int(num) if num.is_integer() else num
                    except ValueError:
                        value = raw
                row_values[idx] = value
            rows.append(row_values)
        return rows


def parse_dt(value: object) -> datetime:
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    return dt.replace(tzinfo=None)


def load_meals(path: str) -> list[dict]:
    rows = read_first_sheet_xlsx(path)
    if not rows:
        raise ValueError("餐食 Excel 为空。")
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}
    required = ["timestamp", "餐次", "食物"]
    missing = [x for x in required if x not in idx]
    if missing:
        raise ValueError(f"餐食表缺少列: {missing}")

    meals = []
    for row in rows[1:]:
        if idx["timestamp"] >= len(row) or not row[idx["timestamp"]]:
            continue
        def get(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None
        meals.append({
            "time": parse_dt(get("timestamp")),
            "meal_type": str(get("餐次") or ""),
            "food": str(get("食物") or ""),
            "exercise": str(get("餐后运动") or ""),
            "notes": str(get("备注") or ""),
        })
    meals.sort(key=lambda x: x["time"])
    return meals


def load_cgm(path: str) -> tuple[list[datetime], np.ndarray]:
    times, values = [], []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Event Type") != "EGV":
                continue
            ts = row.get("Timestamp (YYYY-MM-DDThh:mm:ss)", "").strip()
            gv = row.get("Glucose Value (mg/dL)", "").strip()
            if not ts or not gv:
                continue
            try:
                times.append(parse_dt(ts))
                values.append(float(gv))
            except (ValueError, TypeError):
                continue
    order = np.argsort(np.array(times, dtype="datetime64[us]"))
    times = [times[i] for i in order]
    values = np.asarray(values, dtype=float)[order]
    return times, values


def fmt_num(x, digits=1):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return f"{x:.{digits}f}"


def fmt_metric(x, polluted=False, digits=1, suffix=""):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return f"{x:.{digits}f}{suffix}{' ?' if polluted else ''}"


def coverage_quality(ts: list[datetime], start: datetime, end: datetime) -> tuple[bool, str]:
    if end <= start:
        return False, "窗口长度无效"
    expected = max(1, int((end - start).total_seconds() / 300))
    if len(ts) < max(2, int(expected * 0.70)):
        return False, "CGM覆盖不足"
    gaps = [(b-a).total_seconds()/60 for a, b in zip(ts[:-1], ts[1:])]
    if gaps and max(gaps) > 20:
        return False, f"CGM最大缺口{max(gaps):.0f}分钟"
    return True, ""


def analyze(meals, cgm_times, cgm_values):
    t_arr = cgm_times
    results = []

    def slice_data(start, end):
        lo = bisect_left(t_arr, start)
        hi = bisect_right(t_arr, end)
        return t_arr[lo:hi], cgm_values[lo:hi]

    for i, meal in enumerate(meals):
        mt = meal["time"]
        prev_meal = meals[i-1] if i > 0 else None
        next_meal = meals[i+1] if i+1 < len(meals) else None
        prev_gap = (mt-prev_meal["time"]).total_seconds()/3600 if prev_meal else None
        next_gap = (next_meal["time"]-mt).total_seconds()/3600 if next_meal else None

        # Baseline: median in [-15, 0] minutes. Fallback: nearest within ±15 minutes.
        bts, bvs = slice_data(mt-timedelta(minutes=15), mt)
        baseline_fallback = False
        if len(bvs):
            baseline = float(np.median(bvs))
        else:
            nts, nvs = slice_data(mt-timedelta(minutes=15), mt+timedelta(minutes=15))
            if len(nvs):
                distances = np.array([abs((x-mt).total_seconds()) for x in nts])
                baseline = float(nvs[int(np.argmin(distances))])
                baseline_fallback = True
            else:
                baseline = np.nan

        base_polluted = (prev_gap is not None and prev_gap < 2.0) or baseline_fallback
        reasons = []
        if prev_gap is not None and prev_gap < 2.0:
            reasons.append("上一餐<2h，基线可能受影响")
        if baseline_fallback:
            reasons.append("餐前15分钟无值，基线使用最近值")
        if math.isnan(baseline):
            reasons.append("餐点附近无CGM基线")

        metrics = {}
        pollution = {}

        for hours in (2, 4):
            end = mt + timedelta(hours=hours)
            ts, vs = slice_data(mt, end)
            clean_end = end if next_meal is None else min(end, next_meal["time"])
            clean_ts, _ = slice_data(mt, clean_end)
            overlap = next_gap is not None and next_gap < hours
            complete, msg = coverage_quality(ts, mt, end)
            polluted = base_polluted or overlap or not complete or math.isnan(baseline)
            pollution[f"avg{hours}"] = polluted
            if overlap:
                reasons.append(f"{hours}h窗口与下一餐重叠")
            if not complete and msg:
                reasons.append(f"{hours}h窗口{msg}")
            metrics[f"avg{hours}"] = float(np.mean(vs-baseline)) if len(vs) and not math.isnan(baseline) else np.nan

        for hours in (2, 4):
            end = mt + timedelta(hours=hours)
            ts, vs = slice_data(mt, end)
            overlap = next_gap is not None and next_gap < hours
            complete, msg = coverage_quality(ts, mt, end)
            polluted = base_polluted or overlap or not complete or math.isnan(baseline)
            pollution[f"peak{hours}"] = polluted
            if len(vs) and not math.isnan(baseline):
                j = int(np.argmax(vs))
                metrics[f"peak{hours}"] = float(vs[j]-baseline)
                metrics[f"ttp{hours}"] = (ts[j]-mt).total_seconds()/60
            else:
                metrics[f"peak{hours}"] = np.nan
                metrics[f"ttp{hours}"] = np.nan
            pollution[f"ttp{hours}"] = polluted

        # Half-return time after the 4h peak. Search until 4h or next meal, whichever comes first.
        end4 = mt + timedelta(hours=4)
        search_end = end4 if next_meal is None else min(end4, next_meal["time"])
        ts4, vs4 = slice_data(mt, search_end)
        half_time = np.nan
        half_not_reached = False
        if len(vs4) and not math.isnan(baseline):
            peak_idx = int(np.argmax(vs4))
            peak_inc = float(vs4[peak_idx]-baseline)
            threshold = baseline + peak_inc/2
            for k in range(peak_idx+1, len(vs4)):
                if vs4[k] <= threshold:
                    half_time = (ts4[k]-mt).total_seconds()/60
                    break
            if math.isnan(half_time):
                half_not_reached = True
        metrics["half_return"] = half_time
        half_polluted = base_polluted or (next_gap is not None and next_gap < 4) or half_not_reached
        pollution["half_return"] = half_polluted
        if half_not_reached:
            reasons.append("峰后未在可用窗口降至一半")

        # Remove duplicate reasons while preserving order.
        reasons = list(dict.fromkeys(reasons))
        any_polluted = any(pollution.values())

        result = {
            "餐点时间": mt.strftime("%Y-%m-%d %H:%M"),
            "餐次": meal["meal_type"],
            "食物": meal["food"],
            "基线(mg/dL)": fmt_metric(baseline, base_polluted),
            "2h增量峰值": fmt_metric(metrics["peak2"], pollution["peak2"], suffix=""),
            "2h到峰值(min)": fmt_metric(metrics["ttp2"], pollution["ttp2"], digits=0),
            "4h增量峰值": fmt_metric(metrics["peak4"], pollution["peak4"]),
            "4h到峰值(min)": fmt_metric(metrics["ttp4"], pollution["ttp4"], digits=0),
            "降至一半峰值时间(min)": ("未降至一半 ?" if math.isnan(metrics["half_return"]) and half_not_reached
                                 else fmt_metric(metrics["half_return"], pollution["half_return"], digits=0)),
            "2h平均增量": fmt_metric(metrics["avg2"], pollution["avg2"]),
            "4h平均增量": fmt_metric(metrics["avg4"], pollution["avg4"]),
            "距上一餐(h)": fmt_num(prev_gap, 2),
            "上一餐内容": prev_meal["food"] if prev_meal else "",
            "距下一餐(h)": fmt_num(next_gap, 2),
            "下一餐内容": next_meal["food"] if next_meal else "",
            "污染?": "?" if any_polluted else "",
            "污染原因": "；".join(reasons),
            "_numeric": {
                "2h增量峰值": metrics["peak2"],
                "4h增量峰值": metrics["peak4"],
                "4h到峰值(min)": metrics["ttp4"],
                "降至一半峰值时间(min)": metrics["half_return"],
                "2h平均增量": metrics["avg2"],
                "4h平均增量": metrics["avg4"],
            },
            "_pollution": {
                "2h增量峰值": pollution["peak2"],
                "4h增量峰值": pollution["peak4"],
                "4h到峰值(min)": pollution["ttp4"],
                "降至一半峰值时间(min)": pollution["half_return"],
                "2h平均增量": pollution["avg2"],
                "4h平均增量": pollution["avg4"],
            }
        }
        results.append(result)
    return results


def make_heatmap(results, output_png):
    metric_names = [
        "2h增量峰值", "4h增量峰值", "4h到峰值(min)",
        "降至一半峰值时间(min)", "2h平均增量", "4h平均增量"
    ]
    labels = [f"{r['餐点时间'][5:]} {r['餐次']}\n{r['食物'][:18]}" for r in results]
    data = np.array([[r["_numeric"][m] for m in metric_names] for r in results], dtype=float)

    # Column-wise z-score makes metrics with different units comparable.
    z = np.full_like(data, np.nan)
    for j in range(data.shape[1]):
        col = data[:, j]
        mask = np.isfinite(col)
        if mask.sum() >= 2 and np.nanstd(col) > 0:
            z[mask, j] = (col[mask] - np.nanmean(col)) / np.nanstd(col)
        elif mask.sum():
            z[mask, j] = 0

    height = max(9, 0.36 * len(results) + 2)
    fig, ax = plt.subplots(figsize=(14, height))
    im = ax.imshow(z, aspect="auto", cmap="RdYlGn_r", vmin=-2, vmax=2)
    ax.set_xticks(range(len(metric_names)), metric_names, rotation=35, ha="right")
    ax.set_yticks(range(len(labels)), labels, fontsize=8)
    ax.set_title("餐后血糖指标 Heatmap（颜色为各指标列内 z-score；? 表示可能污染）")

    for i, r in enumerate(results):
        for j, m in enumerate(metric_names):
            val = data[i, j]
            if np.isfinite(val):
                txt = f"{val:.0f}" if "min" in m else f"{val:.1f}"
                if r["_pollution"][m]:
                    txt += "?"
                ax.text(j, i, txt, ha="center", va="center", fontsize=7)
            else:
                ax.text(j, i, "—", ha="center", va="center", fontsize=7)

    fig.colorbar(im, ax=ax, label="列内标准化 z-score")
    fig.tight_layout()
    fig.savefig(output_png, dpi=180, bbox_inches="tight")
    plt.close(fig)


def make_workbook(results, output_xlsx):
    headers = [
        "餐点时间","餐次","食物","基线(mg/dL)",
        "2h增量峰值","2h到峰值(min)","4h增量峰值","4h到峰值(min)",
        "降至一半峰值时间(min)","2h平均增量","4h平均增量",
        "距上一餐(h)","上一餐内容","距下一餐(h)","下一餐内容","污染?","污染原因"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "餐次对比"
    ws.append(headers)
    for r in results:
        ws.append([r[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    widths = {
        "A":17,"B":8,"C":34,"D":13,"E":13,"F":15,"G":13,"H":15,"I":21,
        "J":13,"K":13,"L":13,"M":12,"N":32,"O":12,"P":32,"Q":8,"R":42
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    yellow_font = Font(bold=True, color="9C6500")
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        FormulaRule(formula=['Q2="?"'], fill=yellow_fill, font=yellow_font)
    )
    ws.auto_filter.ref = f"A1:R{ws.max_row}"

    metrics = [
        "2h增量峰值","4h增量峰值","4h到峰值(min)",
        "降至一半峰值时间(min)","2h平均增量","4h平均增量"
    ]
    hs = wb.create_sheet("Heatmap数据")
    hs.append(["餐点时间","餐次","食物"] + metrics)
    for r in results:
        hs.append([
            r["餐点时间"], r["餐次"], r["食物"],
            *[None if not np.isfinite(r["_numeric"][m]) else float(r["_numeric"][m]) for m in metrics]
        ])

    for cell in hs[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    hs.freeze_panes = "A2"
    hs.auto_filter.ref = f"A1:J{hs.max_row}"
    hs.column_dimensions["A"].width = 17
    hs.column_dimensions["B"].width = 8
    hs.column_dimensions["C"].width = 34
    for col in "DEFGHIJ":
        hs.column_dimensions[col].width = 18

    hs.conditional_formatting.add(
        f"D2:J{hs.max_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"
        )
    )

    note = wb.create_sheet("说明")
    note_rows = [
        ["项目","定义"],
        ["基线","优先使用餐前15分钟内CGM中位数；无值时使用餐点±15分钟最近值并标记?。"],
        ["增量峰值","窗口内最高CGM减去基线。"],
        ["平均增量","窗口内所有CGM相对基线的平均值。"],
        ["到峰值时间","从餐点到窗口内最高CGM的分钟数。"],
        ["下降到一半峰值时间","从餐点起算，4h峰值后首次降到 基线 + 峰值增量/2 的时间。"],
        ["污染标记?","上一餐<2h、下一餐进入指标窗口、CGM覆盖不足/缺口>20分钟、使用替代基线，或未降至一半。"],
        ["Heatmap","PNG颜色按每个指标列内z-score标准化；Excel Heatmap数据页使用各列条件色阶。"],
    ]
    for row in note_rows:
        note.append(row)
    for cell in note[1]:
        cell.fill = header_fill
        cell.font = header_font
    note.column_dimensions["A"].width = 24
    note.column_dimensions["B"].width = 90
    for row in note.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_xlsx)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--cgm", required=True)
    p.add_argument("--meals", required=True)
    p.add_argument("--outdir", default="./output")
    args = p.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    meals = load_meals(args.meals)
    cgm_times, cgm_values = load_cgm(args.cgm)
    results = analyze(meals, cgm_times, cgm_values)

    # CSV is convenient for downstream Python/R use.
    csv_out = outdir / "餐后血糖指标对比.csv"
    headers = [k for k in results[0].keys() if not k.startswith("_")]
    with open(csv_out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in results:
            w.writerow({h:r[h] for h in headers})

    make_heatmap(results, outdir / "餐后血糖指标_heatmap.png")
    make_workbook(results, outdir / "餐后血糖指标对比.xlsx")
    print(f"完成：{len(results)} 餐；CGM点数：{len(cgm_times)}")


if __name__ == "__main__":
    main()
