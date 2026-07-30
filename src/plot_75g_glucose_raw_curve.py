from pathlib import Path
from datetime import datetime, timedelta
import csv
from bisect import bisect_left, bisect_right
import argparse

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd

from config import DATA_DIR, DOWNLOADS_DIR
from helper import localize

SCRIPT_VERSION = "v1-75g-glucose-raw-4h"


def setup_chinese_font():
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
    ]
    for font_path in candidates:
        if Path(font_path).exists():
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
            break
    plt.rcParams["axes.unicode_minus"] = False


def parse_dt(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    s = str(value).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    return datetime.fromisoformat(s).replace(tzinfo=None)


def load_meals(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    ts_col = "Meal_Timestamp" if "Meal_Timestamp" in idx else "timestamp"
    food_col = "食物" if "食物" in idx else "Food"
    meals = []
    for row in rows:
        ts = row[idx[ts_col]] if idx.get(ts_col) is not None else None
        if ts is None:
            continue
        food = row[idx[food_col]] if idx.get(food_col) is not None else ""
        meal_type = row[idx["餐次"]] if idx.get("餐次") is not None else ""
        meals.append({
            "time": parse_dt(ts),
            "meal_type": str(meal_type or ""),
            "food": str(food or ""),
        })
    meals.sort(key=lambda x: x["time"])
    return meals


def load_cgm(path):
    times, values = [], []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Event Type") != "EGV":
                continue
            ts = row.get("Timestamp (YYYY-MM-DDThh:mm:ss)", "").strip()
            gv = row.get("Glucose Value (mg/dL)", "").strip()
            if not ts or not gv:
                continue
            try:
                times.append(parse_dt(ts))
                values.append(float(gv))
            except Exception:
                pass
    order = np.argsort(np.array(times, dtype="datetime64[us]"))
    return [times[i] for i in order], np.asarray(values, dtype=float)[order]


def slice_cgm(times, values, start, end):
    lo = bisect_left(times, start)
    hi = bisect_right(times, end)
    return times[lo:hi], values[lo:hi]


def match_75g_glucose(meals):
    candidates = []
    for meal in meals:
        t = meal["food"].lower()
        if "75g glucose" in t or "75 g glucose" in t or ("75g" in t and "glucose" in t) or ("75 g" in t and "glucose" in t):
            candidates.append(meal)
    return candidates


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--meals", default=str(DATA_DIR / "Food_track_202606.xlsx"))
    parser.add_argument("--cgm", default=str(DATA_DIR / "Clarity_Export_Chen_Wei_2026-07-03_145534.csv"))
    parser.add_argument("--outdir", default=str(DOWNLOADS_DIR))
    args = parser.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    setup_chinese_font()

    print(f"Script version: {SCRIPT_VERSION}")

    meals = load_meals(args.meals)
    cgm_times, cgm_values = load_cgm(args.cgm)

    hits = match_75g_glucose(meals)
    if not hits:
        raise SystemExit(localize("未在餐食记录中找到 75g glucose。", "75g glucose not found in meal records."))

    meal = hits[0]
    meal_time = meal["time"]
    ts, vals = slice_cgm(cgm_times, cgm_values, meal_time, meal_time + timedelta(hours=4))

    if len(ts) == 0:
        raise SystemExit(localize("找到 75g glucose，但 4 小时窗口内没有 CGM 数据。", "Found 75g glucose, but no CGM data in 4h window."))

    hours = np.array([(t - meal_time).total_seconds() / 3600 for t in ts], dtype=float)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(hours, vals, linewidth=2, label=meal["food"])
    ax.set_title(localize("75g glucose 的餐后 4 小时血糖曲线", "75g Glucose: 4h Postprandial Blood Glucose Curve"))
    ax.set_xlabel(localize("餐后时间（小时）", "Time after meal (hours)"))
    ax.set_ylabel(localize(f"血糖（mg/dL）", f"Blood Glucose (mg/dL)"))
    ax.set_xlim(0, 4)
    ax.set_xticks([0, 1, 2, 3, 4])
    ax.grid(alpha=0.3)
    ax.legend()
    fig.tight_layout()

    plot_path = outdir / "75g_glucose_4h血糖曲线.png"
    fig.savefig(plot_path, dpi=200, bbox_inches="tight")
    plt.close(fig)

    data_rows = []
    for h, raw in zip(hours, vals):
        data_rows.append({
            localize("匹配时间", "Match Time"): meal_time.strftime("%Y-%m-%d %H:%M"),
            localize("餐次", "Meal Type"): meal["meal_type"],
            localize("食物", "Food"): meal["food"],
            localize("餐后时间(h)", "Time after meal(h)"): round(float(h), 3),
            localize(f"血糖(mg/dL)", f"Blood Glucose(mg/dL)"): round(float(raw), 1),
        })
    pd.DataFrame(data_rows).to_csv(outdir / "75g_glucose_4h血糖曲线_数据.csv", index=False, encoding="utf-8-sig")

    summary = pd.DataFrame([{
        localize("匹配时间", "Match Time"): meal_time.strftime("%Y-%m-%d %H:%M"),
        localize("餐次", "Meal Type"): meal["meal_type"],
        localize("食物", "Food"): meal["food"],
        localize("4小时CGM点数", "CGM points in 4h"): len(ts),
    }])
    summary.to_csv(outdir / "75g_glucose_匹配明细.csv", index=False, encoding="utf-8-sig")

    print(localize("已生成:", "Generated:"))
    print(plot_path)
    print(outdir / "75g_glucose_4h血糖曲线_数据.csv")
    print(outdir / "75g_glucose_匹配明细.csv")


if __name__ == "__main__":
    main()
