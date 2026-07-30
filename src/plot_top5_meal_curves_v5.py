
from __future__ import annotations

import argparse
import csv
import math
from bisect import bisect_left, bisect_right
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from config import UNIT, TEXT_LANGUAGE
from helper import localize

SCRIPT_VERSION = "v5-hours-axis-peak-mean-composite-nextmeal2h"

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请运行: pip install openpyxl") from exc


def parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    s = str(value).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    return datetime.fromisoformat(s).replace(tzinfo=None)


def setup_chinese_font():
    """Use a CJK font when available; otherwise matplotlib falls back automatically."""
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
    ]
    for font_path in candidates:
        if Path(font_path).exists():
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            plt.rcParams["font.family"] = font_manager.FontProperties(
                fname=font_path
            ).get_name()
            break
    plt.rcParams["axes.unicode_minus"] = False


def load_meals(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    required = ["timestamp", "餐次", "食物"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(f"餐食表缺少列: {missing}")

    meals = []
    for row in rows:
        if row[idx["timestamp"]] is None:
            continue

        def get(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        meals.append(
            {
                "time": parse_dt(get("timestamp")),
                "meal_type": str(get("餐次") or ""),
                "food": str(get("食物") or ""),
                "exercise": str(get("餐后运动") or ""),
                "notes": str(get("备注") or ""),
            }
        )
    meals.sort(key=lambda x: x["time"])
    return meals


def load_cgm(path: str) -> tuple[list[datetime], np.ndarray]:
    times, values = [], []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Event Type") != "EGV":
                continue
            ts = row.get("Timestamp (YYYY-MM-DDThh:mm:ss)", "").strip()
            gv = row.get("Glucose Value (mg/dL)", "").strip()
            if not ts or not gv:
                continue
            try:
                times.append(parse_dt(ts))
                values.append(float(gv))
            except (TypeError, ValueError):
                continue

    order = np.argsort(np.array(times, dtype="datetime64[us]"))
    return [times[i] for i in order], np.asarray(values, dtype=float)[order]


def analyze_clean_meals(
    meals: list[dict],
    cgm_times: list[datetime],
    cgm_values: np.ndarray,
    window_hours: float = 4.0,
    min_prev_gap_hours: float = 2.0,
    exclude_keywords: tuple[str, ...] = ("glucose", "葡萄糖"),
) -> list[dict]:
    """
    只保留干净餐：
    - 不限制上一餐间隔
    - 仅排除下一餐出现在当前餐后 min_prev_gap_hours 内的情况
    - 餐前15分钟有CGM基线
    - 4小时窗口覆盖率 >= 70%
    - 最大CGM缺口 <= 20分钟
    """
    records = []

    def get_slice(start: datetime, end: datetime):
        lo = bisect_left(cgm_times, start)
        hi = bisect_right(cgm_times, end)
        return cgm_times[lo:hi], cgm_values[lo:hi]

    for i, meal in enumerate(meals):
        food_lower = meal["food"].lower()
        if any(k.lower() in food_lower for k in exclude_keywords if k):
            continue

        mt = meal["time"]
        prev_gap = (
            (mt - meals[i - 1]["time"]).total_seconds() / 3600
            if i > 0
            else math.inf
        )
        next_gap = (
            (meals[i + 1]["time"] - mt).total_seconds() / 3600
            if i + 1 < len(meals)
            else math.inf
        )

        if next_gap < min_prev_gap_hours:
            continue

        baseline_times, baseline_values = get_slice(
            mt - timedelta(minutes=15), mt
        )
        if len(baseline_values) == 0:
            continue
        baseline = float(np.median(baseline_values))

        end = mt + timedelta(hours=window_hours)
        ts, glucose = get_slice(mt, end)
        expected_points = max(1, int(window_hours * 60 / 5))
        if len(glucose) < int(expected_points * 0.70):
            continue

        gaps = [
            (b - a).total_seconds() / 60
            for a, b in zip(ts[:-1], ts[1:])
        ]
        if gaps and max(gaps) > 20:
            continue

        minutes = np.array(
            [(t - mt).total_seconds() / 60 for t in ts], dtype=float
        )
        increments = glucose - baseline

        # 插值到统一的5分钟网格，便于公平比较和画图。
        grid = np.arange(0, window_hours * 60 + 0.1, 5.0)
        curve = np.interp(grid, minutes, increments)

        peak = float(np.max(curve))
        trough = float(np.min(curve))
        mean_inc = float(np.mean(curve))
        auc_positive = float(np.trapz(np.clip(curve, 0, None), grid) / 60)
        rms_excursion = float(np.sqrt(np.mean(curve ** 2)))
        max_abs_excursion = float(np.max(np.abs(curve)))

        records.append(
            {
                **meal,
                "baseline": baseline,
                "grid": grid,
                "curve": curve,
                "peak": peak,
                "trough": trough,
                "mean_inc": mean_inc,
                "positive_auc": auc_positive,
                "rms_excursion": rms_excursion,
                "max_abs_excursion": max_abs_excursion,
                "prev_gap": prev_gap,
                "next_gap": next_gap,
            }
        )

    return records


def rank_meals(records: list[dict], n: int = 5):
    """
    最糟糕：
    - 4小时增量峰值与4小时平均增幅分别做z-score标准化
    - 两项各占50%，得到综合分
    - 按综合分从高到低排序

    最平稳：
    - 优先按RMS波动
    - 其次按最大绝对偏离
    """
    peaks = np.array([r["peak"] for r in records], dtype=float)
    means = np.array([r["mean_inc"] for r in records], dtype=float)

    peak_std = float(np.std(peaks))
    mean_std = float(np.std(means))

    peak_z = (
        (peaks - float(np.mean(peaks))) / peak_std
        if peak_std > 0 else np.zeros_like(peaks)
    )
    mean_z = (
        (means - float(np.mean(means))) / mean_std
        if mean_std > 0 else np.zeros_like(means)
    )

    for r, pz, mz in zip(records, peak_z, mean_z):
        r["peak_z"] = float(pz)
        r["mean_inc_z"] = float(mz)
        r["worst_score"] = 0.5 * float(pz) + 0.5 * float(mz)

    worst = sorted(
        records,
        key=lambda r: (r["worst_score"], r["mean_inc"], r["peak"]),
        reverse=True,
    )[:n]

    stable = sorted(
        records,
        key=lambda r: (r["rms_excursion"], r["max_abs_excursion"]),
    )[:n]

    return worst, stable



def chinese_food_name(food: str) -> str:
    """Return Chinese food name or English original based on TEXT_LANGUAGE."""
    if TEXT_LANGUAGE == "en":
        return food

    """Translate common English meal names in this dataset into concise Chinese labels."""
    exact = {
        "Kung Pao chicken with white rice": "宫保鸡丁配白米饭",
        "Chow mein, mushroom chicken, Beijing beef": "炒面、蘑菇鸡肉、北京牛肉",
        "Sardine, 300g sweet potato": "沙丁鱼配300克红薯",
        "2 eggs, instant ramen": "两个鸡蛋配方便面",
        "75g pasta with tomato sauce, shrimp, mushrooms": "75克意面配番茄酱、虾和蘑菇",
        "chobani zero sugar yogurt + pistachios": "Chobani零糖酸奶配开心果",
        "Orange, Oikos yogurt, milk": "橙子、Oikos酸奶和牛奶",
        "Pistachio, Oikos Triple Zero yogurt, milk": "开心果、Oikos零糖酸奶和牛奶",
        "1/2 cup oatmeal, milk, protein powder": "半杯燕麦、牛奶和蛋白粉",
        "1/2 cup of oats, milk, protein powder": "半杯燕麦、牛奶和蛋白粉",
        "Pepperoni pizza": "意大利辣香肠披萨",
        "oikos triple zero mixed berry": "Oikos零糖混合莓果酸奶",
    }
    if food in exact:
        return exact[food]

    replacements = [
        ("Kung Pao chicken", "宫保鸡丁"),
        ("white rice", "白米饭"),
        ("Chow mein", "炒面"),
        ("chow mein", "炒面"),
        ("mushroom chicken", "蘑菇鸡肉"),
        ("Beijing beef", "北京牛肉"),
        ("Sardine", "沙丁鱼"),
        ("sardine", "沙丁鱼"),
        ("sweet potato", "红薯"),
        ("instant ramen", "方便面"),
        ("eggs", "鸡蛋"),
        ("egg", "鸡蛋"),
        ("pasta", "意面"),
        ("tomato sauce", "番茄酱"),
        ("shrimp", "虾"),
        ("mushrooms", "蘑菇"),
        ("mushroom", "蘑菇"),
        ("Orange", "橙子"),
        ("orange", "橙子"),
        ("yogurt", "酸奶"),
        ("milk", "牛奶"),
        ("protein powder", "蛋白粉"),
        ("pistachios", "开心果"),
        ("Pistachio", "开心果"),
        ("oatmeal", "燕麦"),
        ("oats", "燕麦"),
        ("cup", "杯"),
    ]
    result = food
    for src, dst in replacements:
        result = result.replace(src, dst)
    return result

def meal_label(record: dict, rank: int) -> str:
    date = record["time"].strftime("%m/%d %H:%M")
    food = chinese_food_name(record["food"].replace("\n", " "))
    if len(food) > 28:
        food = food[:27] + "…"
    return (
        f"{rank}. {date} {record['meal_type']} | {food} "
        f"| {localize('峰值', 'Peak')}+{record['peak']:.0f}，" 
        f"{localize('4h均值', '4h avg')}+{record['mean_inc']:.1f}，"
        f"{localize('综合分', 'Composite score')}{record.get('worst_score', float('nan')):.2f}"
    )


def plot_group(records: list[dict], title: str, output_path: Path):
    fig, ax = plt.subplots(figsize=(12, 7))

    for rank, record in enumerate(records, start=1):
        ax.plot(
            record["grid"],
            record["curve"],
            linewidth=2.2,
            label=meal_label(record, rank),
        )

    ax.axhline(0, linewidth=1)
    ax.axvline(60, linewidth=0.8, linestyle="--", alpha=0.5)
    ax.axvline(120, linewidth=0.8, linestyle="--", alpha=0.5)
    ax.axvline(240, linewidth=0.8, linestyle="--", alpha=0.5)

    ax.set_title(title, fontsize=16)
    ax.set_xlabel(localize("餐后时间（小时）", "Time after meal (hours)"))
    ax.set_ylabel(localize(f"相对餐前基线的血糖增量（{UNIT}）", f"Glucose increase from pre-meal baseline ({UNIT})"))
    ax.set_xlim(0, 240)
    ax.set_xticks([0, 60, 120, 180, 240])
    ax.set_xticklabels(["0", "1", "2", "3", "4"])
    ax.grid(alpha=0.25)
    ax.legend(loc="best", fontsize=9, frameon=True)
    fig.tight_layout()
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def save_selection_csv(worst, stable, output_path: Path):
    fields = [
        localize("分组", "Group"), localize("排名", "Rank"), localize("餐点时间", "Meal Time"), localize("餐次", "Meal Type"), localize("食物", "Food"),
        localize("基线", "Baseline"), localize("4h增量峰值", "4h Peak Increase"), localize("4h平均增量", "4h Avg Increase"), localize("综合排序分", "Composite Score"),
        localize("RMS波动", "RMS Variation"), localize("最大绝对偏离", "Max Absolute Deviation"),
        localize("距上一餐(h)", "Hours Since Prev Meal"), localize("距下一餐(h)", "Hours Until Next Meal"),
    ]
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for group_name, group in [("Top 5 最糟糕", worst), ("Top 5 最平稳", stable)]:
            for rank, r in enumerate(group, start=1):
                writer.writerow(
                    {
                        "分组": group_name,
                        "排名": rank,
                        "餐点时间": r["time"].strftime("%Y-%m-%d %H:%M"),
                        "餐次": r["meal_type"],
                        "食物": chinese_food_name(r["food"]),
                        "基线": round(r["baseline"], 1),
                        "4h增量峰值": round(r["peak"], 1),
                        "4h平均增量": round(r["mean_inc"], 1),
                        "综合排序分": round(r.get("worst_score", float("nan")), 3),
                        "RMS波动": round(r["rms_excursion"], 1),
                        "最大绝对偏离": round(r["max_abs_excursion"], 1),
                        "距上一餐(h)": round(r["prev_gap"], 2),
                        "距下一餐(h)": round(r["next_gap"], 2),
                    }
                )


def main():
    print(f"脚本版本: {SCRIPT_VERSION}")
    parser = argparse.ArgumentParser(
        description="自动选择并绘制Top 5最糟糕餐与Top 5最平稳餐的4小时增量曲线。"
    )
    parser.add_argument("--cgm", default='./Clarity_Export_Chen_Wei_2026-07-03_145534.csv', help="Clarity CGM CSV")
    parser.add_argument("--meals", default='./Stelo_CGM_餐食记录模板.xlsx', help="餐食记录 XLSX")
    parser.add_argument("--outdir", default="./output", help="输出目录")
    parser.add_argument("--top-n", type=int, default=5)
    parser.add_argument(
        "--include-glucose-test",
        action="store_true",
        help="包含75g glucose/葡萄糖测试；默认排除。",
    )
    args = parser.parse_args()

    setup_chinese_font()
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    meals = load_meals(args.meals)
    cgm_times, cgm_values = load_cgm(args.cgm)
    excludes = () if args.include_glucose_test else ("glucose", "葡萄糖")

    records = analyze_clean_meals(
        meals, cgm_times, cgm_values, exclude_keywords=excludes
    )
    if len(records) < args.top_n:
        raise RuntimeError(
            f"只有 {len(records)} 顿满足干净数据标准，少于 top-n={args.top_n}。"
        )

    worst, stable = rank_meals(records, args.top_n)

    plot_group(
        worst,
        localize(f"Top {args.top_n} 最糟糕餐：4小时餐后血糖增量曲线", f"Top {args.top_n} Worst Meals: 4h Postprandial Glucose Increase"),
        outdir / "Top5_最糟糕餐_曲线图.png",
    )
    plot_group(
        stable,
        localize(f"Top {args.top_n} 最平稳餐：4小时餐后血糖增量曲线", f"Top {args.top_n} Most Stable Meals: 4h Postprandial Glucose Increase"),
        outdir / "Top5_最平稳餐_曲线图.png",
    )
    save_selection_csv(
        worst, stable, outdir / "Top5_餐次选择明细.csv"
    )

    print(localize(f"干净餐数量: {len(records)}", f"Clean meals: {len(records)}"))
    print(localize("\nTop 最糟糕餐:", "\nTop Worst Meals:"))
    for i, r in enumerate(worst, 1):
        print(
            f"{i}. {r['time']:%Y-%m-%d %H:%M} | {chinese_food_name(r['food'])} | "
            f"峰值 +{r['peak']:.1f}, 4h平均增幅 {r['mean_inc']:.1f}, 综合分 {r['worst_score']:.2f}"
        )
    print("\nTop 最平稳餐:")
    for i, r in enumerate(stable, 1):
        print(
            f"{i}. {r['time']:%Y-%m-%d %H:%M} | {chinese_food_name(r['food'])} | "
            f"RMS {r['rms_excursion']:.1f}, 最大偏离 {r['max_abs_excursion']:.1f}"
        )


if __name__ == "__main__":
    main()
