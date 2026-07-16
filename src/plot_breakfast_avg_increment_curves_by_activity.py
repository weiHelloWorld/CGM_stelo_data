from pathlib import Path
from datetime import datetime, timedelta
from bisect import bisect_left, bisect_right
import csv
import argparse
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

SCRIPT_VERSION = "v1-breakfast-average-increment-curves-by-activity"

def setup_chinese_font():
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
    ]
    for font_path in candidates:
        if Path(font_path).exists():
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
            break
    plt.rcParams["axes.unicode_minus"] = False

def parse_dt(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    s = str(value).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    return datetime.fromisoformat(s).replace(tzinfo=None)

def parse_duration_to_min(s):
    if not s:
        return np.nan
    parts = str(s).split(":")
    try:
        if len(parts) == 3:
            h, m, sec = map(float, parts)
            return h * 60 + m + sec / 60
    except Exception:
        pass
    return np.nan

def load_meals(meals_path):
    wb = load_workbook(meals_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    required = ["timestamp", "餐次", "食物"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"餐食 Excel 缺少列: {missing}")
    meals = []
    for row in rows:
        ts = row[idx["timestamp"]]
        if ts is None:
            continue
        def get(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None
        meal_time = parse_dt(get("timestamp"))
        if meal_time is None:
            continue
        meals.append({
            "time": meal_time,
            "meal_type": str(get("餐次") or ""),
            "food": str(get("食物") or ""),
        })
    meals.sort(key=lambda x: x["time"])
    return meals

def load_cgm_and_activities(path):
    cgm_times, cgm_values = [], []
    activities = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            event_type = row.get("Event Type")
            ts = parse_dt(row.get("Timestamp (YYYY-MM-DDThh:mm:ss)", ""))
            if event_type == "EGV":
                gv = row.get("Glucose Value (mg/dL)", "").strip()
                if ts is not None and gv:
                    try:
                        cgm_times.append(ts)
                        cgm_values.append(float(gv))
                    except Exception:
                        pass
            elif event_type == "Activity" and ts is not None:
                activities.append({
                    "time": ts,
                    "duration_min": parse_duration_to_min(row.get("Duration (hh:mm:ss)", "")),
                    "raw_duration": row.get("Duration (hh:mm:ss)", ""),
                })
    order = np.argsort(np.array(cgm_times, dtype="datetime64[us]"))
    cgm_times = [cgm_times[i] for i in order]
    cgm_values = np.asarray(cgm_values, dtype=float)[order]
    activities.sort(key=lambda x: x["time"])
    return cgm_times, cgm_values, activities

def slice_cgm(cgm_times, cgm_values, start, end):
    lo = bisect_left(cgm_times, start)
    hi = bisect_right(cgm_times, end)
    return cgm_times[lo:hi], cgm_values[lo:hi]

def baseline_before_meal(cgm_times, cgm_values, meal_time):
    _, bvs = slice_cgm(cgm_times, cgm_values, meal_time - timedelta(minutes=15), meal_time)
    if len(bvs):
        return float(np.median(bvs)), "餐前15分钟中位数"
    nts, nvs = slice_cgm(cgm_times, cgm_values, meal_time - timedelta(minutes=15), meal_time + timedelta(minutes=15))
    if len(nvs):
        distances = np.array([abs((x - meal_time).total_seconds()) for x in nts])
        return float(nvs[int(np.argmin(distances))]), "最近值替代"
    return np.nan, "无基线"

def activities_within_30min(activities, meal_time):
    start = meal_time
    end = meal_time + timedelta(minutes=30)
    return [a for a in activities if start <= a["time"] <= end]

def make_increment_curve(cgm_times, cgm_values, meal_time, baseline, grid_min):
    ts, vs = slice_cgm(cgm_times, cgm_values, meal_time, meal_time + timedelta(hours=4))
    if len(vs) < 2 or not np.isfinite(baseline):
        return None, 0
    minutes = np.array([(t - meal_time).total_seconds() / 60 for t in ts], dtype=float)
    increments = np.asarray(vs, dtype=float) - baseline
    curve = np.full_like(grid_min, np.nan, dtype=float)
    mask = (grid_min >= minutes.min()) & (grid_min <= minutes.max())
    if mask.sum() >= 2:
        curve[mask] = np.interp(grid_min[mask], minutes, increments)
    return curve, len(vs)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--meals", default='./Stelo_CGM_餐食记录模板.xlsx', help="餐食记录 XLSX")
    parser.add_argument("--cgm", default='./Clarity_Export_Chen_Wei_2026-07-03_145534.csv', help="Clarity CGM CSV")
    parser.add_argument("--outdir", default="./output", help="输出目录")
    args = parser.parse_args()

    setup_chinese_font()
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    meals = load_meals(args.meals)
    cgm_times, cgm_values, activities = load_cgm_and_activities(args.cgm)

    grid_min = np.arange(0, 241, 5, dtype=float)
    group_curves = {"有": [], "无": []}
    detail_rows = []

    for meal in meals:
        if meal["meal_type"] != "早餐":
            continue
        meal_time = meal["time"]
        baseline, baseline_source = baseline_before_meal(cgm_times, cgm_values, meal_time)
        acts = activities_within_30min(activities, meal_time)
        group = "有" if acts else "无"
        curve, n_points = make_increment_curve(cgm_times, cgm_values, meal_time, baseline, grid_min)
        if curve is None:
            continue
        group_curves[group].append(curve)
        detail_rows.append({
            "时间": meal_time.strftime("%Y-%m-%d %H:%M"),
            "食物": meal["food"],
            "半小时内有运动": group,
            "运动开始时间": "; ".join(a["time"].strftime("%Y-%m-%d %H:%M:%S") for a in acts),
            "运动持续时间(min)": "; ".join("" if np.isnan(a["duration_min"]) else f'{a["duration_min"]:.1f}' for a in acts),
            "基线(mg/dL)": round(baseline, 1) if np.isfinite(baseline) else np.nan,
            "基线来源": baseline_source,
            "4h窗口CGM点数": n_points,
        })

    fig, ax = plt.subplots(figsize=(11, 7))
    labels = {"有": "半小时内有Activity", "无": "半小时内无Activity"}
    curve_rows = []
    summary_rows = []

    for group in ["有", "无"]:
        curves = np.array(group_curves[group], dtype=float)
        if curves.size == 0:
            continue
        mean_curve = np.nanmean(curves, axis=0)
        std_curve = np.nanstd(curves, axis=0, ddof=1) if curves.shape[0] >= 2 else np.zeros_like(mean_curve)
        count_curve = np.sum(np.isfinite(curves), axis=0)
        hours = grid_min / 60.0

        line, = ax.plot(hours, mean_curve, linewidth=2.5, label=f"{labels[group]} 平均 (n={curves.shape[0]})")
        color = line.get_color()
        ax.fill_between(hours, mean_curve - std_curve, mean_curve + std_curve, color=color, alpha=0.18, label=f"{labels[group]} ±1 std")

        for h, m, s, c in zip(hours, mean_curve, std_curve, count_curve):
            curve_rows.append({
                "分组": labels[group],
                "餐后时间(h)": round(float(h), 3),
                "平均血糖增量(mg/dL)": round(float(m), 3) if np.isfinite(m) else np.nan,
                "std(mg/dL)": round(float(s), 3) if np.isfinite(s) else np.nan,
                "参与均值的曲线数": int(c),
            })

        for target_min in [30, 60, 120, 180, 240]:
            idx = int(np.where(grid_min == target_min)[0][0])
            summary_rows.append({
                "分组": labels[group],
                "n": curves.shape[0],
                "时间点": f"{target_min//60}h" if target_min % 60 == 0 else f"{target_min}min",
                "平均增量": round(float(mean_curve[idx]), 2) if np.isfinite(mean_curve[idx]) else np.nan,
                "std": round(float(std_curve[idx]), 2) if np.isfinite(std_curve[idx]) else np.nan,
                "参与曲线数": int(count_curve[idx]),
            })

    ax.axhline(0, linewidth=1)
    ax.set_title("早餐后4小时平均血糖增量曲线：有运动 vs 无运动", fontsize=20)
    ax.set_xlabel("餐后时间（小时）", fontsize=20)
    ax.set_ylabel("血糖增量（mg/dL，相对餐前基线）", fontsize=20)
    ax.set_xlim(0, 4)
    ax.set_xticks([0, 1, 2, 3, 4])
    ax.grid(alpha=0.3)
    ax.legend(fontsize=16)
    fig.tight_layout()

    plot_path = outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_std阴影.png"
    fig.savefig(plot_path, dpi=200, bbox_inches="tight")
    plt.close(fig)

    pd.DataFrame(curve_rows).to_csv(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_均值std.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(detail_rows).to_csv(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_明细.csv", index=False, encoding="utf-8-sig")
    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_csv(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_摘要.csv", index=False, encoding="utf-8-sig")

    print(f"脚本版本: {SCRIPT_VERSION}")
    print(f"有运动早餐: {len(group_curves['有'])}")
    print(f"无运动早餐: {len(group_curves['无'])}")
    print(summary_df.to_string(index=False))
    print("已生成:")
    print(plot_path)
    print(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_均值std.csv")
    print(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_明细.csv")
    print(outdir / "早餐_有运动_vs无运动_4h平均血糖增量曲线_摘要.csv")

if __name__ == "__main__":
    main()
