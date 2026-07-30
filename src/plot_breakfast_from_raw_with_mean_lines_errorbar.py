from pathlib import Path
from datetime import datetime, timedelta
from bisect import bisect_left, bisect_right
import csv
import argparse

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from config import MG_DL_TO_MMOL_L, UNIT, TEXT_LANGUAGE, DATA_DIR, DOWNLOADS_DIR
from helper import localize

SCRIPT_VERSION = "v1-from-raw-breakfast-activity-mean-lines-errorbar"


def setup_chinese_font():
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
    ]
    for font_path in candidates:
        if Path(font_path).exists():
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
            break
    plt.rcParams["axes.unicode_minus"] = False


def parse_dt(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    s = str(value).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    return datetime.fromisoformat(s).replace(tzinfo=None)


def parse_duration_to_min(s):
    if not s:
        return np.nan
    parts = str(s).split(":")
    try:
        if len(parts) == 3:
            h, m, sec = map(float, parts)
            return h * 60 + m + sec / 60
    except Exception:
        pass
    return np.nan


def load_meals(meals_path):
    wb = load_workbook(meals_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    ts_col = "Meal_Timestamp" if "Meal_Timestamp" in idx else "timestamp"
    meal_col = "餐次"
    food_col = "食物" if "食物" in idx else "Food"
    required = [ts_col, meal_col, food_col]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(localize(f"餐食 Excel 缺少列: {missing}", f"Meal Excel missing columns: {missing}"))

    meals = []
    for row in rows:
        ts = row[idx[ts_col]]
        if ts is None:
            continue

        def get(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        meal_time = parse_dt(get(ts_col))
        if meal_time is None:
            continue

        meals.append({
            "time": meal_time,
            "meal_type": str(get("餐次") or ""),
            "food": str(get(food_col) or ""),
        })

    meals.sort(key=lambda x: x["time"])
    return meals


def load_cgm_and_activities(clarity_csv_path):
    cgm_times, cgm_values = [], []
    activities = []

    with open(clarity_csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            event_type = row.get("Event Type")
            ts = parse_dt(row.get("Timestamp (YYYY-MM-DDThh:mm:ss)", ""))

            if event_type == "EGV":
                gv = row.get("Glucose Value (mg/dL)", "").strip()
                if ts is not None and gv:
                    try:
                        cgm_times.append(ts)
                        value = float(gv)
                        if UNIT == "mmol/L":
                            value = value * MG_DL_TO_MMOL_L
                        cgm_values.append(value)
                    except Exception:
                        pass

            elif event_type == "Activity" and ts is not None:
                activities.append({
                    "time": ts,
                    "duration_min": parse_duration_to_min(row.get("Duration (hh:mm:ss)", "")),
                    "raw_duration": row.get("Duration (hh:mm:ss)", ""),
                })

    if not cgm_times:
        raise ValueError("Clarity CSV 里没有读取到 EGV 血糖数据。")

    order = np.argsort(np.array(cgm_times, dtype="datetime64[us]"))
    cgm_times = [cgm_times[i] for i in order]
    cgm_values = np.asarray(cgm_values, dtype=float)[order]

    activities.sort(key=lambda x: x["time"])
    return cgm_times, cgm_values, activities


def slice_cgm(cgm_times, cgm_values, start, end):
    lo = bisect_left(cgm_times, start)
    hi = bisect_right(cgm_times, end)
    return cgm_times[lo:hi], cgm_values[lo:hi]


def baseline_before_meal(cgm_times, cgm_values, meal_time):
    # 优先用餐前 15 分钟内 CGM 中位数
    _, bvs = slice_cgm(cgm_times, cgm_values, meal_time - timedelta(minutes=15), meal_time)
    if len(bvs):
        return float(np.median(bvs)), "餐前15分钟中位数"

    # 如果餐前没有点，用餐点前后 15 分钟最近值兜底
    nts, nvs = slice_cgm(cgm_times, cgm_values, meal_time - timedelta(minutes=15), meal_time + timedelta(minutes=15))
    if len(nvs):
        distances = np.array([abs((x - meal_time).total_seconds()) for x in nts])
        return float(nvs[int(np.argmin(distances))]), "最近值替代"

    return np.nan, "无基线"


def activities_within_30min(activities, meal_time):
    start = meal_time
    end = meal_time + timedelta(minutes=30)
    return [a for a in activities if start <= a["time"] <= end]


def build_breakfast_metrics(meals, cgm_times, cgm_values, activities):
    records = []

    for meal in meals:
        if meal["meal_type"] != "早餐":
            continue

        meal_time = meal["time"]
        baseline, baseline_source = baseline_before_meal(cgm_times, cgm_values, meal_time)

        ts2, vs2 = slice_cgm(cgm_times, cgm_values, meal_time, meal_time + timedelta(hours=2))
        ts4, vs4 = slice_cgm(cgm_times, cgm_values, meal_time, meal_time + timedelta(hours=4))

        peak2 = float(np.max(vs2) - baseline) if len(vs2) and np.isfinite(baseline) else np.nan
        avg4 = float(np.mean(vs4 - baseline)) if len(vs4) and np.isfinite(baseline) else np.nan

        acts = activities_within_30min(activities, meal_time)

        records.append({
            localize("时间", "Time"): meal_time.strftime("%Y-%m-%d %H:%M"),
            localize("食物", "Food"): meal["food"],
            localize(f"基线({UNIT})", f"Baseline({UNIT})"): round(float(baseline), 1) if np.isfinite(baseline) else np.nan,
            localize("基线来源", "Baseline source"): baseline_source,
            localize("半小时内有运动", "Activity within 30min"): localize("有", "With") if acts else localize("无", "Without"),
            localize("运动开始时间", "Activity start"): "; ".join(a["time"].strftime("%Y-%m-%d %H:%M:%S") for a in acts),
            localize("运动持续时间(min)", "Activity duration(min)"): "; ".join("" if np.isnan(a["duration_min"]) else f'{a["duration_min"]:.1f}' for a in acts),
            localize("4h平均增量", "4h avg increase"): round(avg4, 2) if np.isfinite(avg4) else np.nan,
            localize("2h峰值", "2h peak"): round(peak2, 2) if np.isfinite(peak2) else np.nan,
        })

    return pd.DataFrame(records)


def plot_breakfast_scatter(df, outdir):
    col_4h = localize("4h平均增量", "4h avg increase")
    col_2h = localize("2h峰值", "2h peak")
    plot_df = df.dropna(subset=[col_4h, col_2h]).copy()
    if plot_df.empty:
        raise ValueError(localize("没有可画图的早餐数据：2h峰值或4h平均增量为空。", "No plottable breakfast data: 2h peak or 4h avg increase is empty."))

    fig, ax = plt.subplots(figsize=(11, 8))

    group_order = [
        (localize("有", "With"), localize("半小时内有Activity", "Activity within 30min")),
        (localize("无", "Without"), localize("半小时内无Activity", "No Activity within 30min")),
    ]
    color_map = {}

    # 原始散点
    for key, label in group_order:
        sub = plot_df[plot_df[localize("半小时内有运动", "Activity within 30min")] == key]
        if len(sub) == 0:
            continue

        sc = ax.scatter(
            sub[localize("4h平均增量", "4h avg increase")],
            sub[localize("2h峰值", "2h peak")],
            label=localize(f"{label} 原始点 (n={len(sub)})", f"{label} (n={len(sub)})"),
            s=55,
            alpha=0.75,
        )
        color_map[key] = sc.get_facecolor()[0]

    # Date annotations
    for _, row in plot_df.iterrows():
        ax.annotate(
            pd.to_datetime(row[localize("时间", "Time")]).strftime("%m/%d"),
            (row[col_4h], row[col_2h]),
            fontsize=8,
            xytext=(4, 4),
            textcoords="offset points",
        )

    summary_rows = []

    # Mean dash lines + errorbar at intersection
    for key, label in group_order:
        sub = plot_df[plot_df[localize("半小时内有运动", "Activity within 30min")] == key]
        if len(sub) < 2:
            continue

        x_mean = sub[col_4h].mean()
        x_std = sub[col_4h].std(ddof=1)
        y_mean = sub[col_2h].mean()
        y_std = sub[col_2h].std(ddof=1)
        color = color_map.get(key)

        ax.axvline(
            x_mean,
            linestyle="--",
            linewidth=2,
            alpha=0.85,
            color=color,
            # label=f"{label} 4h均值={x_mean:.1f}",
        )
        ax.axhline(
            y_mean,
            linestyle="--",
            linewidth=2,
            alpha=0.85,
            color=color,
            # label=f"{label} 2h均值={y_mean:.1f}",
        )

        ax.errorbar(
            x_mean,
            y_mean,
            xerr=x_std,
            yerr=y_std,
            fmt="o",
            markersize=12,
            capsize=7,
            capthick=2,
            elinewidth=2.2,
            color=color,
            # label=f"{label} std: x±{x_std:.1f}, y±{y_std:.1f}",
            zorder=5,
        )

        summary_rows.append({
            localize("分组", "Group"): label,
            "n": len(sub),
            localize("4h平均增量_mean", "4h avg increase_mean"): round(x_mean, 2),
            localize("4h平均增量_std", "4h avg increase_std"): round(x_std, 2),
            localize("2h峰值_mean", "2h peak_mean"): round(y_mean, 2),
            localize("2h峰值_std", "2h peak_std"): round(y_std, 2),
        })

    ax.set_title(localize("所有早餐：4h平均增量 vs 2h峰值", "All breakfasts: 4h avg increase vs 2h peak"), fontsize=20)
    ax.set_xlabel(localize(f"4h平均增量（{UNIT}，相对餐前基线）", f"4h avg increase ({UNIT}, relative to pre-meal baseline)"), fontsize=20)
    ax.set_ylabel(localize(f"2h峰值高度（{UNIT}，相对餐前基线）", f"peak ({UNIT}, relative to pre-meal baseline)"), fontsize=20)
    ax.grid(alpha=0.3)
    ax.legend(fontsize=8, loc="best")

    fig.tight_layout()

    plot_path = outdir / "早餐_4h平均增量_vs_2h峰值_从原始文件_均值虚线_errorbar.png"
    fig.savefig(plot_path, dpi=200, bbox_inches="tight")
    plt.close(fig)

    summary_df = pd.DataFrame(summary_rows)
    summary_path = outdir / "早餐_4h平均增量_vs_2h峰值_从原始文件_分组均值std.csv"
    summary_df.to_csv(summary_path, index=False, encoding="utf-8-sig")

    return plot_path, summary_path, summary_df


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--cgm", default=str(DATA_DIR / "Clarity_Export_Chen_Wei_2026-07-03_145534.csv"), help="Clarity CGM CSV")
    parser.add_argument("--meals", default=str(DATA_DIR / "Food_track_202606.xlsx"), help=localize("餐食记录 XLSX", "Meal log XLSX"))
    parser.add_argument("--outdir", default=str(DOWNLOADS_DIR), help=localize("输出目录", "Output directory"))
    args = parser.parse_args()

    setup_chinese_font()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Script version: {SCRIPT_VERSION}")

    meals = load_meals(args.meals)
    cgm_times, cgm_values, activities = load_cgm_and_activities(args.cgm)

    df = build_breakfast_metrics(meals, cgm_times, cgm_values, activities)

    detail_path = outdir / "早餐_4h平均增量_vs_2h峰值_从原始文件_明细.csv"
    df.to_csv(detail_path, index=False, encoding="utf-8-sig")

    plot_path, summary_path, summary_df = plot_breakfast_scatter(df, outdir)

    motion_col = localize("半小时内有运动", "Activity within 30min")
    with_val = localize("有", "With")
    without_val = localize("无", "Without")
    print(localize(f"早餐数量: {len(df)}", f"Breakfast count: {len(df)}"))
    print(localize(f"半小时内有 Activity: {(df[motion_col] == with_val).sum()}", f"With Activity within 30min: {(df[motion_col] == with_val).sum()}"))
    print(localize(f"半小时内无 Activity: {(df[motion_col] == without_val).sum()}", f"Without Activity within 30min: {(df[motion_col] == without_val).sum()}"))
    print()
    print(summary_df.to_string(index=False))
    print()
    print(localize("已生成:", "Generated:"))
    print(plot_path)
    print(detail_path)
    print(summary_path)


if __name__ == "__main__":
    main()
